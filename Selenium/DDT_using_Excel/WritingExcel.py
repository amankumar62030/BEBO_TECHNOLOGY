import os
import openpyxl         #package is required to perform DDT


#File---->Workbook---->sheet---->rows---->cells

file = os.getcwd() + r"\test1.xlsx"

workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet2"]      #or sheet=workbook["Data"]  or sheet=workbook["Sheet1"]


#this will create 5 rows and 3 cols but with same data
#5 rows and 3 cols

for r in range(1,6):
    for c in range(1,4):
        sheet.cell(r,c).value="Welcome"

workbook.save(file)     #in case of writing the file

#create the file with multiple data
#sheet.cell(1,2).value=123
#sheet.cell(1,3).value="Scott"
