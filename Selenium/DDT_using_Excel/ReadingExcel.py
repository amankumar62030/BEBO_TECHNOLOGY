import os
import openpyxl         #package is required to perform DDT


#File---->Workbook---->sheet---->rows---->cells

file = os.getcwd() + r"\test1.xlsx"

workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]

rows = sheet.max_row       #count number of rows in excel
print("Number of rows are: ",rows)

cols = sheet.max_column   #count number of column in excel
print("number of column are: ",cols)

#reading all the rows and column from excel

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value,end="      ")
    print()

